import openpyxl
from openpyxl.utils import get_column_letter

# 株価コード、銘柄名、市場・商品区分、33業種コード・名、17業種コード・名の元データシート読み込み
wb = openpyxl.load_workbook('インサート用データ/data_j.xlsx')

# シート名指定
sheet = wb.get_sheet_by_name('Sheet1')

# 登録対象データカラム検索用リスト
stock_master_data = []

# 全カラム読み込み
for row in sheet['A1':get_column_letter(sheet.max_column) + str(1)]:
    for cell in row:
        stock_master_data.append(str(cell.value))
    print(stock_master_data)

# マスタ登録対象カラム検索
target_columns = []
columns_counter = 0
for column_name in stock_master_data:
    if column_name == 'コード' or \
            column_name == '銘柄名' or \
            column_name == '市場・商品区分' or \
            column_name == '33業種コード' or \
            column_name == '33業種区分' or \
            column_name == '17業種コード' or\
            column_name == '17業種区分':
        target_columns.append(columns_counter)
    columns_counter += 1

# マスタ登録対象データ用変数
list_brand_cd = []                # 株式コード
list_brand_name = []              # 銘柄名
list_brand_type_name = []         # 市場・商品区分
list_industry_code_33_cd = []     # 33業種コード
list_industry_code_33_name = []   # 33業種区分
list_industry_code_17_cd = []     # 17業種コード
list_industry_code_17_name = []   # 17業種区分

# 登録データ読み込み
for brand_cd in list(sheet.columns)[target_columns[0]]:
    if brand_cd.value != '-':
        list_brand_cd.append(brand_cd.value)
    else:
        list_brand_cd.append('0')

for brand_name in list(sheet.columns)[target_columns[1]]:
    if brand_name.value != '-':
        list_brand_name.append(brand_name.value)
    else:
        list_brand_name.append('')

for brand_type_name in list(sheet.columns)[target_columns[2]]:
    if brand_type_name.value != '-':
        list_brand_type_name.append(brand_type_name.value)
    else:
        list_brand_type_name.append('')

for industry_code_33_cd in list(sheet.columns)[target_columns[3]]:
    if industry_code_33_cd.value != '-':
        list_industry_code_33_cd.append(industry_code_33_cd.value)
    else:
        list_industry_code_33_cd.append('0')

for industry_code_33_name in list(sheet.columns)[target_columns[4]]:
    if industry_code_33_name.value != '-':
        list_industry_code_33_name.append(industry_code_33_name.value)
    else:
        list_industry_code_33_name.append('')

for industry_code_17_cd in list(sheet.columns)[target_columns[5]]:
    if industry_code_17_cd.value != '-':
        list_industry_code_17_cd.append(industry_code_17_cd.value)
    else:
        list_industry_code_17_cd.append('0')

for industry_code_17_name in list(sheet.columns)[target_columns[6]]:
    if industry_code_17_name.value != '-':
        list_industry_code_17_name.append(industry_code_17_name.value)
    else:
        list_industry_code_17_name.append('')

# 各マスタ登録ようにセットに変換
set_brand_cd = set(list_brand_cd)                            # 株式コード
set_brand_name = set(list_brand_name)                        # 銘柄名
set_brand_type_name = set(list_brand_type_name)              # 市場・商品区分
set_industry_code_33_cd = set(list_industry_code_33_cd)      # 33業種コード
set_industry_code_33_name = set(list_industry_code_33_name)  # 33業種区分
set_industry_code_17_cd = set(list_industry_code_17_cd)      # 17業種コード
set_industry_code_17_name = set(list_industry_code_17_name)  # 17業種区分


str_insert_worktable_sql = 'insert into stock_brand_wk (' \
                         'brand_wk_cd,' \
                         'brand_wk_name,' \
                         'brand_wk_type,' \
                         'industry_wk_code_33_cd,' \
                         'industry_wk_code_33_name,' \
                         'industry_wk_code_17_cd,' \
                         'industry_wk_code_17_name' \
                         ') values ('

list_insert_worktable_sql = []

for i in range(1,len(list_brand_cd)):
    list_insert_worktable_sql.append(str_insert_worktable_sql +
                                     str(list_brand_cd[i]) + ',' +
                                     '\'' + str(list_brand_name[i]) + '\'' + ',' +
                                     '\'' + str(list_brand_type_name[i]) + '\'' + ',' +
                                     str(list_industry_code_33_cd[i]) + ',' +
                                     '\'' + str(list_industry_code_33_name[i]) + '\'' + ',' +
                                     str(list_industry_code_17_cd[i]) + ',' +
                                     '\'' + str(list_industry_code_17_name[i]) + '\');')

path_w = 'SQL生成/InsertMaster.sql'

with open(path_w, mode='w') as f:
    f.write('\n'.join(list_insert_worktable_sql))